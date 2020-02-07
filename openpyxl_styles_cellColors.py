#! python3
'''
test out grey colors and print out

'gainsboro' # DCDCDC rgb(220,220,220)
'light grey' #D3D3D3 rgb(211,211,211)
'silver' #C0C0C0 rgb(192,192,192)
'dark grey' #A9A9A9 rgb(169,169,169)
'grey' #808080 rgb(128,128,128)
'dim grey' #696969 rgb(105,105,105)
'light slate grey' #778899 rgb(119,136,153)
'slate grey' #708090 rgb(112,128,144)
'dark slate grey' #2F4F4F rgb(47,79,79)
'black' #000000 rgb(0,0,0)
'''
# Initialize
print('\n  Initialize')
import os
import openpyxl
# from openpyxl.workbook import Workbook # what is 'import load_workbook'?
from openpyxl.styles import Alignment, Border, Side, NamedStyle, Font, PatternFill, Protection, GradientFill, Color, colors

# Set directory
# os.chdir('D:') # Todo: Change directory
print('Current Working Directory =', os.getcwd()) # Same result: 
print('Current Working Directory = %s' % os.getcwd())
# create workbook
wb = openpyxl.Workbook()
print('  type(wb) =', type(wb))
print('  sheetnames = ', wb.sheetnames)
# Create Sheets
# sheet_def = wb.create_sheet(title='default', index=0)
sheet_grey = wb.create_sheet(title='grey', index=1)
sheet_red = wb.create_sheet(title='red', index=2)
sheet_green = wb.create_sheet(title='green', index=3)
sheet_blue = wb.create_sheet(title='blue', index=4)
sheet_yellow = wb.create_sheet(title='yellow', index=5)
sheet_orange = wb.create_sheet(title='orange', index=6)
sheet_purple = wb.create_sheet(title='purple', index=7)
print('  wb.sheetnames =',  wb.sheetnames) # 
wb.save('openpyxl_styles_cellColors.xlsx')

wb.active = 0 # Highlight: to change active sheet
sheet_def = wb.active
sheet_def.title = 'default formatting'

# Print Options
print('\n  Print Options')
sheet_grey.print_area = 'A1:I15' # Set print_area
sheet_grey.print_options.horizontalCentered = True
sheet_grey.print_options.verticalCentered = True
# Page margins
sheet_grey.page_margins.left = 0.25
sheet_grey.page_margins.right = 0.25
sheet_grey.page_margins.top = 0.75
sheet_grey.page_margins.bottom = 0.75
sheet_grey.page_margins.header = 0.3
sheet_grey.page_margins.footer = 0.3
# Headers & Footers
sheet_grey.oddHeader.center.text = "&[File]"
sheet_grey.oddHeader.center.size = 20
sheet_grey.oddHeader.center.font = "Tahoma, Bold"
sheet_grey.oddHeader.center.color = "000000" # 
sheet_grey.oddFooter.left.text = "&[Tab]"
sheet_grey.oddFooter.left.size = 12
sheet_grey.oddFooter.left.font = "Tahoma, Bold"
sheet_grey.oddFooter.left.color = "000000" # 
sheet_grey.oddFooter.right.text = "&[Path]&[File]"
sheet_grey.oddFooter.right.size = 6
sheet_grey.oddFooter.right.font = "Tahoma, Bold"
sheet_grey.oddFooter.right.color = "000000"
# Column and Row Dimensions
rows = range(2, 15, 1)
sheet_grey.column_dimensions['A'].width = 20.00
sheet_grey.column_dimensions['C'].width = 20.00
sheet_grey.column_dimensions['D'].width = 20.00
sheet_grey.column_dimensions['B'].width = 40.00
for row in rows:
    sheet_grey.row_dimensions[row].height = 20.00
sheet_grey.row_dimensions[1].height = 40.00

# Cell Values
print('\n  Cell Values')
sheet_grey['A2'].value = 'gainsboro' # DCDCDC rgb(220,220,220)
sheet_grey['A3'].value = 'light grey' # D3D3D3 rgb(211,211,211)
sheet_grey['A4'].value = 'silver' # C0C0C0 rgb(192,192,192)
sheet_grey['A5'].value = 'dark grey' # A9A9A9 rgb(169,169,169)
sheet_grey['A6'].value = 'grey' # 808080 rgb(128,128,128)
sheet_grey['A7'].value = 'dim grey' # 696969 rgb(105,105,105)
sheet_grey['A8'].value = 'light slate grey' # 778899 rgb(119,136,153)
sheet_grey['A9'].value = 'slate grey' # 708090 rgb(112,128,144)
sheet_grey['A10'].value = 'dark slate grey' # 2F4F4F rgb(47,79,79)
sheet_grey['A11'].value = 'black' # 000000 rgb(0,0,0)

sheet_grey['C2'].value = '#DCDCDC'
sheet_grey['C3'].value = '#D3D3D3'
sheet_grey['C4'].value = '#C0C0C0'
sheet_grey['C5'].value = '#A9A9A9'
sheet_grey['C6'].value = '#808080'
sheet_grey['C7'].value = '#696969'
sheet_grey['C8'].value = '#778899'
sheet_grey['C9'].value = '#708090'
sheet_grey['C10'].value = '#2F4F4F'
sheet_grey['C11'].value = '#000000'

sheet_grey['D2'].value = 'rgb(220,220,220)'
sheet_grey['D3'].value = 'rgb(211,211,211)'
sheet_grey['D4'].value = 'rgb(192,192,192)'
sheet_grey['D5'].value = 'rgb(169,169,169)'
sheet_grey['D6'].value = 'rgb(128,128,128)'
sheet_grey['D7'].value = 'rgb(105,105,105)'
sheet_grey['D8'].value = 'rgb(119,136,153)'
sheet_grey['D9'].value = 'rgb(112,128,144)'
sheet_grey['D10'].value = 'rgb(47,79,79)'
sheet_grey['D11'].value = 'rgb(0,0,0)'

# Cell Color
print('\n  Cell Color')
rowsA2 = [2]
rowsA3 = [3]
rowsA4 = [4]
rowsA5 = [5]
rowsA6 = [6]
rowsA7 = [7]
rowsA8 = [8]
rowsA9 = [9]
rowsA10 = [10]
rowsA11 = [11]
columns = [2]

'gainsboro' # DCDCDC rgb(220,220,220)
'light grey' #D3D3D3 rgb(211,211,211)
'silver' #C0C0C0 rgb(192,192,192)
'dark grey' #A9A9A9 rgb(169,169,169)
'grey' #808080 rgb(128,128,128)
'dim grey' #696969 rgb(105,105,105)
'light slate grey' #778899 rgb(119,136,153)
'slate grey' #708090 rgb(112,128,144)
'dark slate grey' #2F4F4F rgb(47,79,79)
'black' #000000 rgb(0,0,0)

for col in columns:
    for row in rowsA2:
        sheet_grey.cell(row=row, column=col).fill = PatternFill(fgColor='DCDCDC', fill_type = 'solid')

for col in columns:
    for row in rowsA3:
        sheet_grey.cell(row=row, column=col).fill = PatternFill(fgColor='D3D3D3', fill_type = 'solid')

for col in columns:
    for row in rowsA4:
        sheet_grey.cell(row=row, column=col).fill = PatternFill(fgColor='C0C0C0', fill_type = 'solid')

for col in columns:
    for row in rowsA5:
        sheet_grey.cell(row=row, column=col).fill = PatternFill(fgColor='A9A9A9', fill_type = 'solid')

for col in columns:
    for row in rowsA6:
        sheet_grey.cell(row=row, column=col).fill = PatternFill(fgColor='808080', fill_type = 'solid')

for col in columns:
    for row in rowsA7:
        sheet_grey.cell(row=row, column=col).fill = PatternFill(fgColor='696969', fill_type = 'solid')

for col in columns:
    for row in rowsA8:
        sheet_grey.cell(row=row, column=col).fill = PatternFill(fgColor='778899', fill_type = 'solid')

for col in columns:
    for row in rowsA9:
        sheet_grey.cell(row=row, column=col).fill = PatternFill(fgColor='708090', fill_type = 'solid')

for col in columns:
    for row in rowsA10:
        sheet_grey.cell(row=row, column=col).fill = PatternFill(fgColor='2F4F4F', fill_type = 'solid')

for col in columns:
    for row in rowsA11:
        sheet_grey.cell(row=row, column=col).fill = PatternFill(fgColor='000000', fill_type = 'solid')

wb.save('openpyxl_styles_cellColors.xlsx')
print('\n  openpyxl_styles_cell colors.py complete')