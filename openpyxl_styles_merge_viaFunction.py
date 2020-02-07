#! python

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment

align = Alignment(horizontal='center', vertical='center', wrap_text=True)

def main():
    wb = Workbook()
    filename = 'openpyxl_styles_merge_viaFunction.xlsx'
    wb.active.title = 'merge_example'
    sheet = wb.active
    print('Active sheet is :', sheet)
    wb.save(filename = 'openpyxl_styles_merge_viaFunction.xlsx')
    
    # #1) merge columns in same row
    sheet.merge_cells('A1:F1')
    sheet.cell(row=1, column=1).value = '#1)  merge row 1, columns 3-6'
    sheet.cell(row=1, column=1).alignment = align

    # #2) merge rows in same column
    sheet.merge_cells(start_row=3, start_column=1, end_row=12, end_column=1)
    sheet.cell(row=3, column=1).value = '#2)  merge rows 2 through 11'
    sheet.cell(row=3, column=1).alignment = align

    # #3) merge columns and rows
    sheet.merge_cells(start_row=3, start_column=3, end_row=11, end_column=6)
    sheet.cell(row=3, column=3).value = '#3)  merge row 1-5 & column 3-5'
    sheet.cell(row=3, column=3).alignment = align
    
    wb.save(filename = 'openpyxl_styles_merge_viaFunction.xlsx')

if __name__ == '__main__':
    main()

print('openpyxl_styles_merge_viaFunction.py complete')