# ERRORS: 
# get_sheet_by_name('sheetName') -- replaced by -- wb['Sheet']

# Imports
import os # 
import openpyxl # 
from openpyxl.utils import get_column_letter, column_index_from_string

print('cwd = ', os.getcwd())
# os.chdir ('c:\\users\pc\desktop') # CHANGE DIRECTORY FOR FILE LOCATION

wb = openpyxl.Workbook()
print('  Created a new Excel Spreadsheet')
print ("New spreadsheet type = ", type (wb)) # 
print('all sheetnames = ', wb.sheetnames)
wb.create_sheet()
wb.create_sheet()
wb.create_sheet()
print('all sheetnames = ', wb.sheetnames)

# Saving an Excel Spreadsheet
print('  save the empty file')
wb.save('testFile_cells_00.xlsx') # 

sheet = wb.active
print('\nwb.active = ', sheet)

# Use cell name, as a string, to access cell-object. 
#   Note: Cell objects can contain value, row, column, coordinates'
print('\n  Use cell name, as a string, to access cell-object')
print('    Note: Cell objects can contain value, row, column, coordinates')

# Creating variables containing cell objects
print('  Creating variables containing cell objects')
cellB2 = sheet['B2'] # string ""
cellB3 = sheet['B3'] # integer whole number
cellB4 = sheet['B4'] # float decimal number
cellB5 = sheet['B5'] # boolean ture or false

# Adding data into cells
print('\n  Adding data into cells')
sheet['A1'] = 'Type'
sheet['B1'] = 'Value'
sheet['A2'] = 'String'
sheet['A3'] = 'Integer'
sheet['A4'] = 'Float'
sheet['A5'] = 'Boolean'
sheet['A6'] = 'Formulas:'
sheet['B6'] = 200
sheet['C6'] = 300
sheet['D6'] = 'sum(B6:C6)'
sheet['E6'] = 'Formula should result in cellD6 = 500'

# Cell B2_String
print('  Cell B2_String')
print('cell_B2.value = ', cellB2.value)
print('type(cellB2.value) = ', type(cellB2.value))
cellB2.value = 'A String'
print('  Cell value added\ncell_B2.value = ', cellB2.value)
print('type(cellB2.value) = ', type(cellB2.value))

# Cell B3_Integer
print('\n  Cell B3_Integer')
print('cell_B3.value = ', cellB3.value)
print('type(cellB3.value) = ', type(cellB3.value))
cellB3.value = 1001
print('  Cell value added\ncell_B3.value = ', cellB3.value)
print('type(cellB3.value) = ', type(cellB3.value))

# Cell B4_Float
print('\n  Cell B4_Float')
print('cell_B4.value = ', cellB4.value)
print('type(cellB4.value) = ', type(cellB4.value))
cellB4.value = 100.001
print('  Cell value added\ncell_B4.value = ', cellB4.value)
print('type(cellB4.value) = ', type(cellB4.value))

# Cell B5_Boolean
print('\n  Cell B5_Boolean')
print('cell_B5.value = ', cellB5.value)
print('type(cellB5.value) = ', type(cellB5.value))
cellB5.value = True
print('  Cell value added\ncell_B5.value = ', cellB5.value)
print('type(cellB5.value) = ', type(cellB5.value))

# Saving an Excel Spreadsheet
print('  save the file')
wb.save('testFile_cells_01.xlsx') # 

# Python, by default, returns the cell-value depending on the spreadsheet-cell-formatting (i.e.: date formatting returns a datetime object, etc.)

# Converted a cell.value to a string object (as opposed to the spreadsheet cell format)
print('\n  Converted a cell.value to a string object (as opposed to the spreadsheet cell format)')
print('type(cellB3.value) = ', type(cellB3.value))
print('type(str(cellB3.value)) = ', type(str(cellB3.value))) # str(cellA1.value)
print('  or another way')
print("type(str(sheet01['B3'].value)) = ", type(str(sheet['B3'].value))) # str(cellA1.value)


# Specify row and column values specifically without using cell names
print('\n  Specify row and column values specifically without using cell names')
# Note: Cells & Rows start at 1, not 0
print('sheet.cell(row=1, column=2) = ', sheet.cell(row=1, column=2), '\n') # calling worksheet cell method

# Numbers-to-letters AND Letters-to-Numbers
print('get_column_letter(1) = ', get_column_letter(1))
print('get_column_letter(2) = ', get_column_letter(2))
print('get_column_letter(20) = ', get_column_letter(20))
print('get_column_letter(26) = ', get_column_letter(26))
print('get_column_letter(27) = ', get_column_letter(27))
print('get_column_letter(sheet.max_column) = ', get_column_letter(sheet.max_column))
print("column_index_from_string('A') = ", column_index_from_string('A'))
print("column_index_from_string('B') = ", column_index_from_string('B'))
print("column_index_from_string('T') = ", column_index_from_string('T'))
print("column_index_from_string('Z') = ", column_index_from_string('Z'))
print("column_index_from_string('AA') = ", column_index_from_string('AA'))

# Concatinating Cell Data
print('\nConcatinating Cell Data (Note: Cannot concatinate non-string values(verify this))')
print('  type(cellB2.value) = ', type(cellB2.value))
print("Row " + str(cellB2.row) + ", Column " + str(cellB2.column) + ' value = "' + cellB2.value + '"')
print('Cell ' + cellB2.coordinate + ' value = "' + cellB2.value + '"')
print('  type(cellB3.value) = ', type(cellB3.value))
print("Row " + str(cellB3.row) + ", Column " + str(cellB3.column) + ' value = "' + str(cellB3.value) + '"')
print('Cell ' + cellB3.coordinate + ' value = "' + str(cellB3.value) + '"')



# Looping
print('\n\n  Looping,')
print('          Looping,')
print('                  & Looping\n')

print("\ntuple(sheet['A1':'C3']) = \n", tuple(sheet['A1':'C3']), 'HOW TO GET VALUES, NOT COORDINATES?\n') # HOW TO GET VALUE, NOT COORDINATES?

# forLoop accessing defined row numbers & column values
print('\n   forLoop accessing defined row numbers & column values')
# Note: range is where to start, where to end but must go one over last row &/or column needed
for i in range(1,7): # rows
    for c in range(1,3): # columns
        print(i, sheet.cell(row=i, column=c).value)

# forLoop accessing a defined cell range giving cell names & values
print('\n  forLoop accessing a defined cell range giving cell names & values')
# forLoop: 1st line iterates each row in slice, 2nd line iterates each cell
for rowOfCellObjects in sheet['A1':'B6']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)

# Code below gotten from:  stackoverflow.com/questions/42974450/iterate-over-worksheets-rows-columns

# WTF!! See below
print('\n  Iteration over all rows and columns in one Worksheet')
print('WHY, THIS LOOP WORKS UP HERE,...')
for row_cells in sheet.iter_rows():
    for cell in row_cells:
       print('%s cell.value = %s' % (cell.coordinate, cell.value) )

# Iteration over all columns of one row (i.e. row==2)
print("\n  Iteration over all columns of one row (i.e. row==2)")
for row_cells in sheet.iter_rows(min_row=2, max_row=2):
    for cell in row_cells:
        print('%s cell value = %s' % (cell.coordinate, cell.value) )  

# Iteration over all rows of one column (i.e. column==2)
print('\n  Iteration over all rows of one column (i.e. column==2)')
for col_cells in sheet.iter_cols(min_col=2, max_col=2):
    for cell in col_cells:
        #print('%s: cell.value = %s' % (cell, cell.value))
        print('%s: cell value = %s' % (cell.coordinate, cell.value))

#  Iteration over all worksheets in a workbook
print('\n  Iteration over all worksheets in a workbook')
for n, sheet in enumerate(wb.worksheets):
    print('Sheet Index:[{}], Title:{}'.format(n, sheet.title))

# WTF!!
# Iteration over all rows and columns in one Worksheet
print('\n  Iteration over all rows and columns in one Worksheet')
print('...BUT NOT DOWN HERE')
for row_cells in sheet.iter_rows():
    for cell in row_cells:
       print('%s cell.value = %s' % (cell.coordinate, cell.value) )

# Formulas
print('\n  Formulas')
