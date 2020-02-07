#! python3
print('\n  spreadsheet_default.py started')
'''
Purpose of this file:
# Create Spreadsheet
# Create Sheets
# Enter Cell Values
# Set Font & Cell Color
# Merge Cells
# Add Print Options
# Column and Row Dimensions
# Add Borders
'''
# Create Spreadsheet
print('\n  Create Spreadsheet')
import os
import openpyxl
from openpyxl.styles import Alignment, Border, Side, NamedStyle, Font, PatternFill, Protection, GradientFill, Color, colors
# Set directory
# os.chdir('D:') # Todo: Change directory
print('Current Working Directory = %s' % os.getcwd())
# create workbook
wb = openpyxl.Workbook()
print('type(wb) =', type(wb))

# Create Sheets
print('\n  Create Sheets')
print('sheetnames = ', wb.sheetnames)
# wb.active = 0 # sheet index number
sheet01 = wb.active
sheet01.title = 'sheet01'
sheet02 = wb.create_sheet(title='sheet02', index=1)
sheet03 = wb.create_sheet(title='sheet03', index=2)
print('sheetnames =',  wb.sheetnames) # 
wb.save('_spreadsheet_template.xlsx')


# Cell Values (Note: Both ways enter values into cells)
print('\n  Cell Values')
sheet01['A1'] = ''
sheet01['B1'].value = ''
sheet01['C1'] = ''
sheet01['D1'].value = ''
sheet01['E1'] = ''
sheet01['F1'].value = ''
sheet01['G1'] = ''
sheet01['H1'].value = ''
sheet01['I1'] = ''
sheet01['J1'].value = ''
sheet01['A2'] = 'merged A2:A5'

# Set Font & Cell Color
print('\n  Set Font & Cell Color')
headerFont = Font(name='Calibri', size=12, bold=True, underline='single', italic=False)
for row_cells in sheet01.iter_rows(min_row=1, max_row=1):
    for cell in row_cells:
        cell.font = headerFont
        cell.fill = PatternFill(fgColor='A9A9A9', fill_type = 'solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)

# Merge Cells
print('\n  Merge Cells')
sheet01.merge_cells('A2:A5')
# sheet.unmerge_cells('A1:A5')

# Freeze Panes # See notes at end
print('\n  Freeze Panes')
sheet01.freeze_panes = 'A2'

# Print Options
print('\n  Print Options')
sheet01.print_area = 'A1:J40' # Set print_area
sheet01.print_options.horizontalCentered = True
sheet01.print_options.verticalCentered = True
# Page margins
sheet01.page_margins.left = 0.25
sheet01.page_margins.right = 0.25
sheet01.page_margins.top = 0.75
sheet01.page_margins.bottom = 0.75
sheet01.page_margins.header = 0.3
sheet01.page_margins.footer = 0.3
# Headers & Footers
sheet01.oddHeader.center.text = "&[File]"
sheet01.oddHeader.center.size = 20
sheet01.oddHeader.center.font = "Tahoma, Bold"
sheet01.oddHeader.center.color = "000000" # 
sheet01.oddFooter.left.text = "&[Tab]"
sheet01.oddFooter.left.size = 12
sheet01.oddFooter.left.font = "Tahoma, Bold"
sheet01.oddFooter.left.color = "000000" # 
sheet01.oddFooter.right.text = "&[Path]&[File]"
sheet01.oddFooter.right.size = 6
sheet01.oddFooter.right.font = "Tahoma, Bold"
sheet01.oddFooter.right.color = "000000"

# Column and Row Dimensions
rows = range(2, 20, 1)
cols = range(1, 10, 1)
for row in rows:
    for col in cols:
        sheet01.row_dimensions[row].height = 15.00
        # DELETE: sheet01.row_dimensions[row].width = 200.00
        sheet01.column_dimensions['A'].width = 22.00
        # ERROR: sheet01.column_dimensions[cols].width = 30.00
sheet01.row_dimensions[1].height = 40.00 # header row

# Cell Borders
print('\n  Cell Borders')
borderThick = Border(left=Side(style='thick'), 
                    right=Side(style='thick'), 
                    top=Side(style='thick'), 
                    bottom=Side(style='thick')) #
borderThin = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')) # 
borderTop = Border(top=Side(style='thick'),
                    left=Side(style='thin'),
                    right=Side(style='thin'), 
                    bottom=Side(style='thin'))
borderBottom = Border(bottom=Side(style='thick'),
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'))
borderLeft = Border(left=Side(style='thick'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')) # 
borderRight = Border(right=Side(style='thick'),
                    left=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')) # 
borderTopLeft = Border(left=Side(style='thick'),
                    right=Side(style='thin'),
                    top=Side(style='thick'),
                    bottom=Side(style='thin')) #
borderTopRight = Border(left=Side(style='thin'),
                    right=Side(style='thick'),
                    top=Side(style='thick'),
                    bottom=Side(style='thin')) #
borderBottomLeft = Border(left=Side(style='thick'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thick')) #
borderBottomRight = Border(left=Side(style='thin'),
                    right=Side(style='thick'),
                    top=Side(style='thin'),
                    bottom=Side(style='thick')) #
# Set thin borders on entire sheet
rows = range(1, 41)
rowTop = int(1)
rowBot = int(40)
colLeft = int(1)
colRight = int(10)
columns = range(1, 11)
for row in rows:
    for col in columns:
        sheet01.cell(row, col).border = borderThin
# Set thick outer borders
for row in rows:
    for col in columns:
        sheet01.cell(rowTop, col).border = borderTop
        sheet01.cell(rowBot, col).border = borderBottom
        sheet01.cell(row, colLeft).border = borderLeft
        sheet01.cell(row, colRight).border = borderRight
        sheet01['A1'].border = borderTopLeft
        sheet01['J1'].border = borderTopRight
        sheet01['A40'].border = borderBottomLeft
        sheet01['J40'].border = borderBottomRight

wb.save('_spreadsheet_template.xlsx')
print('\n  spreadsheet_default.py complete\n')