#! python3
# python_excel_Automate the Boring Stuff_Ch12_
'''
# Font() Function Keyword Arguement Attributes:
#   Keyword         Data       Description
    Argument        Type
================================================================
1)   'name'         string      font name (i.e.: 'Times New Roman')
2)   'size'         integer     point size
3)   'bold'         boolean     True for Bold
3)   'underline'    string     'single'
4)   'italic'       boolean     True for Italic 
'''
import openpyxl
from openpyxl.styles import Font
# create workbook
wb = openpyxl.Workbook()
print('  sheetnames = ', wb.sheetnames)

sheet = wb['Sheet']
# Enter cell values
sheet['A1'] = "Font Type"
sheet['A2'] = "Font size"
sheet['A3'] = "Bold"
sheet['A4'] = "Underline"
sheet['A5'] = "Italic"
sheet['A6'] = "Color"

italic24Font = Font(size=24, italic=True)
sheet['A1'].font = italic24Font

wb.save('setting_cell_font_and_styles.xlsx')