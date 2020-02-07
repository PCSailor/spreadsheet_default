#! python
'''
Border options:
'dashed', 'dotted', 'double', 'hair', 'thin', 'thick', 'mediumDashDot', 'slantDashDot', 'medium', 'mediumDashDotDot', 'dashDot', 'dashDotDot', 'mediumDashed'
'''
# Imports
import os # 
import openpyxl # 
from openpyxl.styles import Border, Side# Alignment

print('cwd = ', os.getcwd())
# os.chdir ('c:\\users\pc\desktop') # CHANGE DIRECTORY FOR FILE LOCATION

wb = openpyxl.Workbook()
print('  Created a new Excel Spreadsheet')
print ("New spreadsheet type = ", type (wb)) # 
print('all sheetnames = ', wb.sheetnames)
sheet = wb.active
print('sheet = ', sheet)

borderThick = Border(left=Side(style='thick'), 
                    right=Side(style='thick'), 
                    top=Side(style='thick'), 
                    bottom=Side(style='thick')) #

borderThin = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')) # 

borderTop = Border(top=Side(style='thick'),
                    left=Side(style='thin'),
                    right=Side(style='thin'), 
                    bottom=Side(style='thin'))

borderBottom = Border(bottom=Side(style='thick'),
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'))

borderLeft = Border(left=Side(style='thick'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')) # 
                    
borderRight = Border(right=Side(style='thick'),
                    left=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')) # 

borderTopLeft = Border(left=Side(style='thick'),
                    right=Side(style='thin'),
                    top=Side(style='thick'),
                    bottom=Side(style='thin')) #

borderTopRight = Border(left=Side(style='thin'),
                    right=Side(style='thick'),
                    top=Side(style='thick'),
                    bottom=Side(style='thin')) #

borderBottomLeft = Border(left=Side(style='thick'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thick')) #

borderBottomRight = Border(left=Side(style='thin'),
                    right=Side(style='thick'),
                    top=Side(style='thin'),
                    bottom=Side(style='thick')) #

borderSide = Side(border_style='thick')  # applies to sides of each cell

# Set thin borders on entire sheet
rows = range(1, 46)
rowTop = int(1)
rowBot = int(45)
colLeft = int(1)
colRight = int(10)
columns = range(1, 11)
for row in rows:
    for col in columns:
        sheet.cell(row, col).border = borderThin

# Set thick outer borders
for row in rows:
    for col in columns:
        sheet.cell(rowTop, col).border = borderTop
        sheet.cell(rowBot, col).border = borderBottom
        sheet.cell(row, colLeft).border = borderLeft
        sheet.cell(row, colRight).border = borderRight
        sheet['A1'].border = borderTopLeft
        sheet['J1'].border = borderTopRight
        sheet['A45'].border = borderBottomLeft
        sheet['J45'].border = borderBottomRight

# Saving an Excel Spreadsheet
print('  save the empty file')
wb.save('openpyxl_styles_cellBorders.xlsx') # 