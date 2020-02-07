# ERRORS: Line 103, 147
# get_sheet_by_name('sheetName') -- replaced by -- wb['Sheet']
# get_column_letter -- replaced by -- 
# column_index_from_string -- replaced by -- 

# Imports
import os # 
import openpyxl # 

print('cwd = ', os.getcwd())
# os.chdir ('c:\\users\pc\desktop') # CHANGE DIRECTORY FOR FILE LOCATION

wb = openpyxl.Workbook()
print('  Created a new Excel Spreadsheet')
print ("  Newly-created spreadsheet type (equals =) ", type (wb)) # 
print('all sheetnames = ', wb.sheetnames)

# Saving an Excel Spreadsheet
print('  save the file')
wb.save('testFile_sheets_00.xlsx') # 

# since new workbook, should result in True
print("wb['Sheet']['A1'].value == None = ", wb['Sheet']['A1'].value == None)

# Get the active sheet title
sheet = wb.active
print('\n  Get the active sheet title')
print('sheet.title = ', sheet.title)
# Retitling Sheets
print('  Retitling Sheets')
print('all sheetnames = ', wb.sheetnames)
print('sheet.title = ', sheet.title)
sheet.title = "sheet_01" # here's the code
print('sheet.title = ', sheet.title)
print('all sheetnames = ', wb.sheetnames)

# Creating Sheets
print('\n  Creating Sheets')
wb.create_sheet('sheet_02') # here's the code
wb.create_sheet('sheet_03')
wb.create_sheet('sheet_04')
print('all sheetnames = ', wb.sheetnames)

# save the file
print('  save the file')
wb.save('testFile_sheets_01.xlsx') #

# Removing Sheets (Note: Change from older code)
print('\n  Removing Sheets')
print('all sheetnames = ', wb.sheetnames)
del wb['sheet_04']
print('all sheetnames after deleting sheet_04 = ', wb.sheetnames)
del wb['sheet_03']
print('all sheetnames after deleting sheet_03 = ', wb.sheetnames)
print('all sheetnames = ', wb.sheetnames)

# create a variable with an object containing a sheet
sheet02 = wb['sheet_02'] # OLD: .get_sheet_by_name('sheet_02')
# check the sheet object using type() function
print('\n  check the sheet object (using type() function)')
print('type(sheet(active one)) = ', type(sheet)) # 
print('type(sheet02) = ', type(sheet02)) # 

'''
# CORRECTION:  Renaming a sheet by index number (How to do this??)
print('  CORRECTION: Renaming a sheet by index number (How to do this??)')
wb.create_sheet(index=1, title='sheetName_02')
print('  wb.sheetnames = (INCORRECT)', wb.sheetnames)
wb.create_sheet(index=2, title='sheetName_03')
print('  wb.sheetnames = (INCORRECT)', wb.sheetnames)

How to change an individual sheet?
wb.sheet[3].title = 'sheetName_04'
change this name with the Worksheet.title property:
ws.title = "New Title"
'''

# save the file
print('  save the file')
wb.save('testFile_sheets_02.xlsx') #
