# ERRORS: Line 
# get_sheet_by_name('sheetName') -- replaced by -- wb['Sheet']
# get_column_letter -- replaced by -- 
# column_index_from_string -- replaced by -- 

# Imports
# With imports, no error messages indicate a successful install
import openpyxl # 
import os # 
# from openpyxl.cell import get_column_letter, column_index_from_string
# import pyinputplus as pyip

print('cwd = ', os.getcwd())
# os.chdir ('c:\\users\pc\desktop') # CHANGE DIRECTORY FOR FILE LOCATION
print(input('Continue or change directory?'))
# pyip.inputYesNo()

''' WORKBOOKS '''
print("\n ''' WORKBOOKS ''' \n")

# Creating a new Excel Spreadsheet
print('  Creating a new Excel Spreadsheet')
# create new workbook/spreadsheet
# create a new, blank Excel ss by calling the Workbook() function (Note: capital 'W')
print("  create a new, blank Excel ss by calling the Workbook() function (Note: capital 'W')")
mtss = openpyxl.Workbook()
print ("Open newly created 'mtss' type = ", type (mtss)) # 
print('  mtss.sheetnames = ', mtss.sheetnames)

# Saving an Excel Spreadsheet
print('  save the file')
mtss.save('testFile_workbooks_00.xlsx') # 


# Opening an existing workbook
print('\n  Opening an existing workbook')
# create an object containing an entire spreadsheet-workbook by opening a .xlsx file
# Question: Will this open openoffice & libreoffice the same way??
wb = openpyxl.load_workbook('testFile_existing_00.xlsx') # CHANGE FILE NAME
print ("Opened 'wb' type = ", type (wb)) # 

# Saving an Excel Spreadsheet
print('  save the file')
wb.save('testFile_existing_01.xlsx') # 